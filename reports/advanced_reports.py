# نظام التقارير المتطور والشامل
# Advanced Comprehensive Reporting System with AI Analytics

import io
import json
import datetime
from decimal import Decimal
from typing import Dict, List, Optional, Any, Union
import base64
import logging

# Django imports
from django.http import HttpResponse
from django.template.loader import get_template
from django.utils import timezone
from django.db.models import Q, Count, Avg, Sum, Max, Min
from django.contrib.auth import get_user_model

# Report generation libraries
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Excel generation
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

# Data analysis
import pandas as pd
import numpy as np

# Visualization
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_agg import FigureCanvasAgg
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend

# Arabic text support
try:
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display
    ARABIC_SUPPORT = True
except ImportError:
    ARABIC_SUPPORT = False

logger = logging.getLogger(__name__)
User = get_user_model()

class ReportGenerator:
    """مولد التقارير المتطور"""
    
    def __init__(self):
        self.setup_fonts()
        self.styles = self._create_styles()
    
    def setup_fonts(self):
        """إعداد الخطوط للنصوص العربية"""
        try:
            # محاولة تسجيل خط عربي
            font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
            pdfmetrics.registerFont(TTFont('Arabic', font_path))
            self.arabic_font = 'Arabic'
        except:
            logger.warning("فشل في تحميل الخط العربي - استخدام الخط الافتراضي")
            self.arabic_font = 'Helvetica'
    
    def _create_styles(self):
        """إنشاء أنماط التقارير"""
        styles = getSampleStyleSheet()
        
        # نمط العنوان الرئيسي
        styles.add(ParagraphStyle(
            name='ArabicTitle',
            parent=styles['Title'],
            fontName=self.arabic_font,
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20,
            textColor=colors.darkblue
        ))
        
        # نمط العنوان الفرعي
        styles.add(ParagraphStyle(
            name='ArabicHeading',
            parent=styles['Heading1'],
            fontName=self.arabic_font,
            fontSize=14,
            alignment=TA_RIGHT,
            spaceAfter=12,
            textColor=colors.darkgreen
        ))
        
        # نمط النص العادي
        styles.add(ParagraphStyle(
            name='ArabicNormal',
            parent=styles['Normal'],
            fontName=self.arabic_font,
            fontSize=10,
            alignment=TA_RIGHT,
            spaceBefore=6,
            spaceAfter=6
        ))
        
        return styles
    
    def _format_arabic_text(self, text: str) -> str:
        """تنسيق النص العربي للعرض الصحيح"""
        if not ARABIC_SUPPORT or not text:
            return text
        
        try:
            reshaped_text = reshape(text)
            return get_display(reshaped_text)
        except:
            return text

class AcademicReportGenerator(ReportGenerator):
    """مولد التقارير الأكاديمية"""
    
    def generate_student_transcript(self, student_id: str, format_type: str = 'pdf') -> Dict:
        """إنشاء كشف درجات الطالب"""
        try:
            from students.models import User, StudentProfile
            from academic.models import StudentGrade, CourseEnrollment
            
            # الحصول على بيانات الطالب
            student = User.objects.get(student_id=student_id, role='STUDENT')
            profile = StudentProfile.objects.get(user=student)
            
            # الحصول على الدرجات
            grades = StudentGrade.objects.filter(
                student=student
            ).select_related('course', 'semester').order_by('-semester__year', 'course__name')
            
            # إنشاء التقرير
            if format_type.lower() == 'pdf':
                return self._create_transcript_pdf(student, profile, grades)
            elif format_type.lower() == 'excel':
                return self._create_transcript_excel(student, profile, grades)
            else:
                raise ValueError("نوع التقرير غير مدعوم")
                
        except Exception as e:
            logger.error(f"خطأ في إنشاء كشف الدرجات: {str(e)}")
            raise
    
    def _create_transcript_pdf(self, student, profile, grades) -> Dict:
        """إنشاء كشف درجات PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=inch, leftMargin=inch)
        story = []
        
        # العنوان الرئيسي
        title = self._format_arabic_text("كشف درجات الطالب")
        story.append(Paragraph(title, self.styles['ArabicTitle']))
        story.append(Spacer(1, 20))
        
        # معلومات الطالب
        student_info = [
            [self._format_arabic_text("رقم الطالب:"), student.student_id],
            [self._format_arabic_text("الاسم:"), self._format_arabic_text(student.display_name)],
            [self._format_arabic_text("التخصص:"), self._format_arabic_text(profile.major.name_ar if profile.major else "غير محدد")],
            [self._format_arabic_text("المعدل التراكمي:"), f"{profile.cumulative_gpa:.3f}"],
            [self._format_arabic_text("الساعات المكتسبة:"), str(profile.completed_credit_hours)],
            [self._format_arabic_text("تاريخ التقرير:"), timezone.now().strftime("%Y-%m-%d")]
        ]
        
        student_table = Table(student_info, colWidths=[2*inch, 3*inch])
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), self.arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(student_table)
        story.append(Spacer(1, 30))
        
        # جدول الدرجات
        if grades:
            # عنوان الجدول
            grades_title = self._format_arabic_text("سجل الدرجات")
            story.append(Paragraph(grades_title, self.styles['ArabicHeading']))
            story.append(Spacer(1, 10))
            
            # رأس الجدول
            header = [
                self._format_arabic_text("المقرر"),
                self._format_arabic_text("الرمز"),
                self._format_arabic_text("الساعات"),
                self._format_arabic_text("الدرجة"),
                self._format_arabic_text("التقدير"),
                self._format_arabic_text("الفصل")
            ]
            
            data = [header]
            
            # إضافة الدرجات
            for grade in grades:
                row = [
                    self._format_arabic_text(grade.course.name_ar),
                    grade.course.code,
                    str(grade.course.credit_hours),
                    f"{grade.numerical_grade:.1f}" if grade.numerical_grade else "غير محدد",
                    grade.letter_grade or "غير محدد",
                    self._format_arabic_text(f"{grade.semester.name} {grade.semester.year}")
                ]
                data.append(row)
            
            grades_table = Table(data, colWidths=[2*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.5*inch])
            grades_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), self.arabic_font),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
            ]))
            
            story.append(grades_table)
        
        # بناء الوثيقة
        doc.build(story)
        buffer.seek(0)
        
        return {
            'success': True,
            'content': base64.b64encode(buffer.getvalue()).decode(),
            'filename': f"transcript_{student.student_id}_{timezone.now().strftime('%Y%m%d')}.pdf",
            'content_type': 'application/pdf'
        }
    
    def _create_transcript_excel(self, student, profile, grades) -> Dict:
        """إنشاء كشف درجات Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "كشف الدرجات"
        
        # تنسيق العنوان
        ws['A1'] = "كشف درجات الطالب"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        # معلومات الطالب
        row = 3
        student_info = [
            ("رقم الطالب:", student.student_id),
            ("الاسم:", student.display_name),
            ("التخصص:", profile.major.name_ar if profile.major else "غير محدد"),
            ("المعدل التراكمي:", f"{profile.cumulative_gpa:.3f}"),
            ("الساعات المكتسبة:", str(profile.completed_credit_hours))
        ]
        
        for label, value in student_info:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # رأس جدول الدرجات
        row = 9
        headers = ["المقرر", "الرمز", "الساعات", "الدرجة", "التقدير", "الفصل"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # بيانات الدرجات
        row += 1
        for grade in grades:
            ws.cell(row=row, column=1, value=grade.course.name_ar)
            ws.cell(row=row, column=2, value=grade.course.code)
            ws.cell(row=row, column=3, value=grade.course.credit_hours)
            ws.cell(row=row, column=4, value=float(grade.numerical_grade) if grade.numerical_grade else 0)
            ws.cell(row=row, column=5, value=grade.letter_grade or "غير محدد")
            ws.cell(row=row, column=6, value=f"{grade.semester.name} {grade.semester.year}")
            row += 1
        
        # تنسيق الأعمدة
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 20
        
        # حفظ الملف
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return {
            'success': True,
            'content': base64.b64encode(buffer.getvalue()).decode(),
            'filename': f"transcript_{student.student_id}_{timezone.now().strftime('%Y%m%d')}.xlsx",
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    
    def generate_course_performance_report(self, course_id: str, semester_id: str) -> Dict:
        """تقرير أداء المقرر"""
        try:
            from courses.models import Course
            from academic.models import StudentGrade, Semester
            
            course = Course.objects.get(id=course_id)
            semester = Semester.objects.get(id=semester_id)
            
            # الحصول على درجات المقرر
            grades = StudentGrade.objects.filter(
                course=course,
                semester=semester
            ).select_related('student')
            
            if not grades:
                return {'success': False, 'error': 'لا توجد درجات لهذا المقرر'}
            
            # تحليل البيانات
            analysis = self._analyze_course_performance(grades, course, semester)
            
            # إنشاء التقرير
            return self._create_performance_report_pdf(analysis)
            
        except Exception as e:
            logger.error(f"خطأ في تقرير أداء المقرر: {str(e)}")
            raise
    
    def _analyze_course_performance(self, grades, course, semester) -> Dict:
        """تحليل أداء المقرر"""
        grade_values = [float(g.numerical_grade) for g in grades if g.numerical_grade]
        
        if not grade_values:
            return {'error': 'لا توجد درجات رقمية للتحليل'}
        
        analysis = {
            'course': course,
            'semester': semester,
            'total_students': len(grades),
            'average_grade': np.mean(grade_values),
            'median_grade': np.median(grade_values),
            'std_deviation': np.std(grade_values),
            'min_grade': min(grade_values),
            'max_grade': max(grade_values),
            'pass_rate': len([g for g in grade_values if g >= 60]) / len(grade_values) * 100,
            'grade_distribution': self._calculate_grade_distribution(grade_values),
            'statistical_summary': self._generate_statistical_summary(grade_values)
        }
        
        return analysis
    
    def _calculate_grade_distribution(self, grades: List[float]) -> Dict:
        """حساب توزيع الدرجات"""
        ranges = {
            'A (90-100)': len([g for g in grades if 90 <= g <= 100]),
            'B (80-89)': len([g for g in grades if 80 <= g < 90]),
            'C (70-79)': len([g for g in grades if 70 <= g < 80]),
            'D (60-69)': len([g for g in grades if 60 <= g < 70]),
            'F (0-59)': len([g for g in grades if g < 60])
        }
        return ranges
    
    def _generate_statistical_summary(self, grades: List[float]) -> Dict:
        """إنشاء ملخص إحصائي"""
        quartiles = np.percentile(grades, [25, 50, 75])
        
        return {
            'first_quartile': quartiles[0],
            'median': quartiles[1],
            'third_quartile': quartiles[2],
            'iqr': quartiles[2] - quartiles[0],
            'outliers': self._detect_outliers(grades, quartiles[0], quartiles[2])
        }
    
    def _detect_outliers(self, grades: List[float], q1: float, q3: float) -> List[float]:
        """اكتشاف القيم الشاذة"""
        iqr = q3 - q1
        lower_bound = q1 - 1.5 * iqr
        upper_bound = q3 + 1.5 * iqr
        
        return [g for g in grades if g < lower_bound or g > upper_bound]

class FinancialReportGenerator(ReportGenerator):
    """مولد التقارير المالية"""
    
    def generate_student_financial_statement(self, student_id: str, academic_year: str = None) -> Dict:
        """بيان مالي للطالب"""
        try:
            from students.models import User
            from finance.models import StudentAccount, Payment, Fee
            
            student = User.objects.get(student_id=student_id, role='STUDENT')
            account = StudentAccount.objects.get(student=student)
            
            # تحديد السنة الأكاديمية
            if not academic_year:
                academic_year = timezone.now().year
            
            # الحصول على البيانات المالية
            payments = Payment.objects.filter(
                student_account=account,
                payment_date__year=academic_year
            ).order_by('-payment_date')
            
            fees = Fee.objects.filter(
                student_accounts=account,
                academic_year=academic_year
            )
            
            # حساب الإجماليات
            total_fees = sum(fee.amount for fee in fees)
            total_payments = sum(payment.amount for payment in payments if payment.status == 'COMPLETED')
            balance = total_fees - total_payments
            
            # إنشاء التقرير
            return self._create_financial_statement_pdf(
                student, account, payments, fees, total_fees, total_payments, balance, academic_year
            )
            
        except Exception as e:
            logger.error(f"خطأ في البيان المالي: {str(e)}")
            raise
    
    def _create_financial_statement_pdf(self, student, account, payments, fees, 
                                      total_fees, total_payments, balance, academic_year) -> Dict:
        """إنشاء البيان المالي PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # العنوان
        title = self._format_arabic_text(f"البيان المالي - العام الأكاديمي {academic_year}")
        story.append(Paragraph(title, self.styles['ArabicTitle']))
        story.append(Spacer(1, 20))
        
        # معلومات الطالب
        student_info = [
            [self._format_arabic_text("رقم الطالب:"), student.student_id],
            [self._format_arabic_text("الاسم:"), self._format_arabic_text(student.display_name)],
            [self._format_arabic_text("رقم الحساب:"), account.account_number],
            [self._format_arabic_text("تاريخ التقرير:"), timezone.now().strftime("%Y-%m-%d")]
        ]
        
        student_table = Table(student_info)
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), self.arabic_font),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT')
        ]))
        
        story.append(student_table)
        story.append(Spacer(1, 30))
        
        # ملخص مالي
        summary_title = self._format_arabic_text("الملخص المالي")
        story.append(Paragraph(summary_title, self.styles['ArabicHeading']))
        
        summary_data = [
            [self._format_arabic_text("إجمالي الرسوم:"), f"{total_fees:,.2f} ريال"],
            [self._format_arabic_text("إجمالي المدفوعات:"), f"{total_payments:,.2f} ريال"],
            [self._format_arabic_text("الرصيد المتبقي:"), f"{balance:,.2f} ريال"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), self.arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT')
        ]))
        
        story.append(summary_table)
        story.append(PageBreak())
        
        # تفاصيل المدفوعات
        if payments:
            payments_title = self._format_arabic_text("تفاصيل المدفوعات")
            story.append(Paragraph(payments_title, self.styles['ArabicHeading']))
            
            payment_headers = [
                self._format_arabic_text("التاريخ"),
                self._format_arabic_text("المبلغ"),
                self._format_arabic_text("النوع"),
                self._format_arabic_text("الحالة"),
                self._format_arabic_text("المرجع")
            ]
            
            payment_data = [payment_headers]
            
            for payment in payments:
                row = [
                    payment.payment_date.strftime("%Y-%m-%d"),
                    f"{payment.amount:,.2f}",
                    self._format_arabic_text(payment.get_payment_type_display()),
                    self._format_arabic_text(payment.get_status_display()),
                    payment.reference_number or "غير محدد"
                ]
                payment_data.append(row)
            
            payment_table = Table(payment_data)
            payment_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), self.arabic_font),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER')
            ]))
            
            story.append(payment_table)
        
        doc.build(story)
        buffer.seek(0)
        
        return {
            'success': True,
            'content': base64.b64encode(buffer.getvalue()).decode(),
            'filename': f"financial_statement_{student.student_id}_{academic_year}.pdf",
            'content_type': 'application/pdf'
        }

class AnalyticsReportGenerator(ReportGenerator):
    """مولد تقارير التحليلات والإحصائيات"""
    
    def generate_enrollment_analytics(self, semester_id: str = None) -> Dict:
        """تحليلات التسجيل"""
        try:
            from academic.models import CourseEnrollment, Semester
            from courses.models import Course, College
            
            # تحديد الفصل الدراسي
            if semester_id:
                semester = Semester.objects.get(id=semester_id)
                enrollments = CourseEnrollment.objects.filter(semester=semester)
            else:
                current_semester = Semester.objects.filter(is_current=True).first()
                enrollments = CourseEnrollment.objects.filter(semester=current_semester)
                semester = current_semester
            
            # تحليل البيانات
            analytics = self._analyze_enrollment_data(enrollments, semester)
            
            # إنشاء الرسوم البيانية
            charts = self._create_enrollment_charts(analytics)
            
            # إنشاء التقرير
            return self._create_analytics_report_pdf(analytics, charts, "تحليلات التسجيل")
            
        except Exception as e:
            logger.error(f"خطأ في تحليلات التسجيل: {str(e)}")
            raise
    
    def _analyze_enrollment_data(self, enrollments, semester) -> Dict:
        """تحليل بيانات التسجيل"""
        from django.db.models import Count
        
        total_enrollments = enrollments.count()
        
        # تحليل حسب الكلية
        college_stats = enrollments.values(
            'course__department__college__name_ar'
        ).annotate(
            count=Count('id')
        ).order_by('-count')
        
        # تحليل حسب المستوى الأكاديمي
        level_stats = enrollments.values(
            'student__studentprofile__academic_level'
        ).annotate(
            count=Count('id')
        ).order_by('student__studentprofile__academic_level')
        
        # تحليل حسب حالة التسجيل
        status_stats = enrollments.values('status').annotate(
            count=Count('id')
        )
        
        return {
            'semester': semester,
            'total_enrollments': total_enrollments,
            'college_distribution': list(college_stats),
            'level_distribution': list(level_stats),
            'status_distribution': list(status_stats),
            'average_enrollments_per_course': enrollments.values('course').distinct().count(),
            'analysis_date': timezone.now()
        }
    
    def _create_enrollment_charts(self, analytics: Dict) -> Dict:
        """إنشاء الرسوم البيانية للتسجيل"""
        charts = {}
        
        # رسم بياني للتوزيع حسب الكلية
        if analytics['college_distribution']:
            plt.figure(figsize=(10, 6))
            colleges = [item['course__department__college__name_ar'] or 'غير محدد' 
                       for item in analytics['college_distribution']]
            counts = [item['count'] for item in analytics['college_distribution']]
            
            plt.bar(range(len(colleges)), counts)
            plt.xlabel('الكليات')
            plt.ylabel('عدد التسجيلات')
            plt.title('توزيع التسجيلات حسب الكلية')
            plt.xticks(range(len(colleges)), colleges, rotation=45, ha='right')
            plt.tight_layout()
            
            # حفظ الرسم
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            buffer.seek(0)
            charts['college_chart'] = base64.b64encode(buffer.getvalue()).decode()
            plt.close()
        
        # رسم دائري للتوزيع حسب المستوى
        if analytics['level_distribution']:
            plt.figure(figsize=(8, 8))
            levels = [f"المستوى {item['student__studentprofile__academic_level']}" 
                     for item in analytics['level_distribution']]
            counts = [item['count'] for item in analytics['level_distribution']]
            
            plt.pie(counts, labels=levels, autopct='%1.1f%%', startangle=90)
            plt.title('توزيع التسجيلات حسب المستوى الأكاديمي')
            plt.axis('equal')
            
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            buffer.seek(0)
            charts['level_chart'] = base64.b64encode(buffer.getvalue()).decode()
            plt.close()
        
        return charts
    
    def _create_analytics_report_pdf(self, analytics: Dict, charts: Dict, title: str) -> Dict:
        """إنشاء تقرير التحليلات PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # العنوان
        report_title = self._format_arabic_text(title)
        story.append(Paragraph(report_title, self.styles['ArabicTitle']))
        story.append(Spacer(1, 20))
        
        # معلومات عامة
        general_info = [
            [self._format_arabic_text("الفصل الدراسي:"), 
             self._format_arabic_text(f"{analytics['semester'].name} {analytics['semester'].year}")],
            [self._format_arabic_text("إجمالي التسجيلات:"), str(analytics['total_enrollments'])],
            [self._format_arabic_text("تاريخ التقرير:"), 
             analytics['analysis_date'].strftime("%Y-%m-%d %H:%M")]
        ]
        
        info_table = Table(general_info)
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), self.arabic_font),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT')
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 30))
        
        # إضافة الرسوم البيانية
        for chart_name, chart_data in charts.items():
            from reportlab.platypus import Image
            
            chart_buffer = io.BytesIO(base64.b64decode(chart_data))
            img = Image(chart_buffer, width=6*inch, height=4*inch)
            story.append(img)
            story.append(Spacer(1, 20))
        
        doc.build(story)
        buffer.seek(0)
        
        return {
            'success': True,
            'content': base64.b64encode(buffer.getvalue()).decode(),
            'filename': f"analytics_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            'content_type': 'application/pdf'
        }

# دوال مساعدة للاستخدام الخارجي

def generate_student_report(student_id: str, report_type: str, format_type: str = 'pdf') -> Dict:
    """إنشاء تقرير طالب"""
    if report_type == 'transcript':
        generator = AcademicReportGenerator()
        return generator.generate_student_transcript(student_id, format_type)
    elif report_type == 'financial':
        generator = FinancialReportGenerator()
        return generator.generate_student_financial_statement(student_id)
    else:
        raise ValueError("نوع التقرير غير مدعوم")

def generate_course_report(course_id: str, semester_id: str, report_type: str = 'performance') -> Dict:
    """إنشاء تقرير مقرر"""
    generator = AcademicReportGenerator()
    return generator.generate_course_performance_report(course_id, semester_id)

def generate_analytics_report(report_type: str, **kwargs) -> Dict:
    """إنشاء تقرير تحليلي"""
    generator = AnalyticsReportGenerator()
    
    if report_type == 'enrollment':
        return generator.generate_enrollment_analytics(kwargs.get('semester_id'))
    else:
        raise ValueError("نوع التحليل غير مدعوم")

def export_data_to_excel(data: List[Dict], filename: str, sheet_name: str = 'البيانات') -> Dict:
    """تصدير البيانات إلى Excel"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            return {'success': False, 'error': 'لا توجد بيانات للتصدير'}
        
        # إضافة الرؤوس
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        
        # إضافة البيانات
        for row, item in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item.get(header, ''))
        
        # تنسيق الأعمدة
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # حفظ الملف
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return {
            'success': True,
            'content': base64.b64encode(buffer.getvalue()).decode(),
            'filename': filename,
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
    except Exception as e:
        logger.error(f"خطأ في تصدير Excel: {str(e)}")
        return {'success': False, 'error': str(e)}